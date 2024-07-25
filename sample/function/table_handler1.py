import sys
import pandas as pd
from PySide6.QtWidgets import QMainWindow, QInputDialog,QVBoxLayout,QColorDialog, QWidget, QTableWidget, QColorDialog,QTableWidgetItem, QApplication, QPushButton, QHBoxLayout, QFileDialog, QMessageBox
from PySide6.QtGui import QColor, QFont
from PySide6.QtCore import Qt
from server.client import MongoClient
from function.table import TableWidget
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time

class TableHandler:
    def __init__(self, table_widget, db_handler):
        self.table_widget = table_widget
        self.db_handler = db_handler

    def save_data(self):
        start_time = time.time()
        table = self.table_widget.get_table()

        # Step 1: Collect data from table
        step_start_time = time.time()
        data = []
        for row in range(table.rowCount()):
            row_data = {}
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    font = item.font()
                    row_data[str(col)] = {
                        'text': item.text(),
                        'foreground': item.foreground().color().name(),
                        'background': item.background().color().name(),
                        'alignment': item.textAlignment(),
                        'font': {
                            'bold': font.bold(),
                            'size': font.pointSize()
                        },
                        'row_height': table.rowHeight(row),
                        'column_width': table.columnWidth(col)
                    }
                else:
                    row_data[str(col)] = {
                        'text': '',
                        'foreground': QColor(Qt.black).name(),
                        'background': QColor(Qt.white).name(),
                        'alignment': int(Qt.AlignLeft | Qt.AlignVCenter),
                        'font': {
                            'bold': False,
                            'size': 10
                        },
                        'row_height': table.rowHeight(row),
                        'column_width': table.columnWidth(col)
                    }
            data.append(row_data)
        step_end_time = time.time()
        print(f"Step 1 (Collect data from table) took {step_end_time - step_start_time:.4f} seconds")

        # Step 2: Collect merged cells information
        step_start_time = time.time()
        merged_cells = []
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                if table.rowSpan(row, col) > 1 or table.columnSpan(row, col) > 1:
                    merged_cells.append({
                        'row': row,
                        'col': col,
                        'row_span': table.rowSpan(row, col),
                        'col_span': table.columnSpan(row, col)
                    })
        step_end_time = time.time()
        print(f"Step 2 (Collect merged cells information) took {step_end_time - step_start_time:.4f} seconds")

        # Step 3: Save data to database
        step_start_time = time.time()
        self.db_handler.save_table(data)
        step_end_time = time.time()
        print(f"Step 3 (Save data to database) took {step_end_time - step_start_time:.4f} seconds")

        # Step 4: Save merged cells to database
        step_start_time = time.time()
        self.db_handler.save_merged_cells(merged_cells)
        step_end_time = time.time()
        print(f"Step 4 (Save merged cells to database) took {step_end_time - step_start_time:.4f} seconds")

        # Step 5: Show confirmation message
        step_start_time = time.time()
        QMessageBox.information(self.table_widget, "保存成功", "表格数据已保存到数据库")
        step_end_time = time.time()
        print(f"Step 5 (Show confirmation message) took {step_end_time - step_start_time:.4f} seconds")

        end_time = time.time()
        print(f"Total save_data execution time: {end_time - start_time:.4f} seconds")

    def refresh_data(self):
        table = self.table_widget.get_table()
        if table.rowCount() > 0:
            table.removeRow(table.rowCount() - 1)
        self.load_table_data()
        QMessageBox.information(self.table_widget, "刷新成功", "表格数据已刷新")


    def export_to_excel(self):
        table = self.table_widget.get_table()
        path, _ = QFileDialog.getSaveFileName(self.table_widget, "导出为Excel", "新建表格.xlsx", "Excel Files (*.xlsx)")
        if not path:  # 如果用户没有选择文件名，则使用默认文件名
            path = "新建表格.xlsx"

        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    cell = ws.cell(row=row + 1, column=col + 1, value=item.text())

                    # 设置对齐方式
                    alignment = item.textAlignment()
                    align = Alignment(
                        horizontal='left' if alignment & Qt.AlignLeft else
                        'right' if alignment & Qt.AlignRight else
                        'center' if alignment & Qt.AlignHCenter else 'general',
                        vertical='top' if alignment & Qt.AlignTop else
                        'bottom' if alignment & Qt.AlignBottom else
                        'center' if alignment & Qt.AlignVCenter else 'center'
                    )
                    cell.alignment = align

                    # 设置字体
                    font = item.font()
                    cell.font = Font(
                        bold=font.bold(),
                        size=font.pointSize()
                    )

                    # 设置前景色和背景色
                    fg_color = item.foreground().color().name()[1:]  # Remove '#' from color code
                    bg_color = item.background().color().name()[1:]
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

                    # 设置单元格大小
                    ws.row_dimensions[row + 1].height = table.rowHeight(row)/2
                    ws.column_dimensions[get_column_letter(col + 1)].width = table.columnWidth(col)/10

                    # 设置边框（默认为Excel边框样式）
                    cell.border = thin_border

        # 合并单元格
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                if table.rowSpan(row, col) > 1 or table.columnSpan(row, col) > 1:
                    ws.merge_cells(
                        start_row=row + 1, start_column=col + 1,
                        end_row=row + table.rowSpan(row, col),
                        end_column=col + table.columnSpan(row, col)
                    )

        wb.save(path)
        QMessageBox.information(self.table_widget, "导出成功", f"表格已成功导出到 {path}")


    def load_table_data(self):
        response = self.db_handler.get_table()
        if response["status"] == "success":
            data = response["data"] # Convert JSON string to dictionary
            self.populate_table(data)
        else:
            self.populate_table_with_default_data()

    def populate_table(self, data):
        start_time = time.time()
        table = self.table_widget.get_table()

        # Step 1: Clear contents
        step_start_time = time.time()
        table.clearContents()
        step_end_time = time.time()
        print(f"Step 1 (Clear contents) took {step_end_time - step_start_time:.4f} seconds")

        # Step 2: Check data format
        step_start_time = time.time()
        if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
            raise ValueError("Data must be a list of dictionaries")
        step_end_time = time.time()
        print(f"Step 2 (Check data format) took {step_end_time - step_start_time:.4f} seconds")

        # Step 3: Set row and column counts
        step_start_time = time.time()
        row_count = len(data)
        col_count = max(len(row) for row in data) if row_count > 0 else 0
        table.setRowCount(row_count)
        table.setColumnCount(col_count)
        step_end_time = time.time()
        print(f"Step 3 (Set row and column counts) took {step_end_time - step_start_time:.4f} seconds")

        # Step 4: Populate table with data
        step_start_time = time.time()
        for row_idx, row_data in enumerate(data):
            if not isinstance(row_data, dict):
                raise ValueError(f"Row data at index {row_idx} is not a dictionary: {row_data}")

            for col_idx_str, cell_data in row_data.items():
                if col_idx_str == 'merged_cells':  # Skip 'merged_cells' key
                    continue

                try:
                    col_idx = int(col_idx_str)  # Convert column index back to integer
                except ValueError:
                    continue  # Skip keys that cannot be converted to integers

                if isinstance(cell_data, dict):
                    item = QTableWidgetItem(cell_data.get('text', ''))
                    item.setForeground(QColor(cell_data.get('foreground', QColor(Qt.black).name())))
                    item.setBackground(QColor(cell_data.get('background', QColor(Qt.white).name())))
                    item.setTextAlignment(cell_data.get('alignment', int(Qt.AlignLeft | Qt.AlignVCenter)))
                    font = item.font()
                    font.setBold(cell_data.get('font', {}).get('bold', False))
                    font.setPointSize(cell_data.get('font', {}).get('size', 10))
                    item.setFont(font)
                    table.setItem(row_idx, col_idx, item)
                    table.setRowHeight(row_idx, cell_data.get('row_height', table.rowHeight(row_idx)))
                    table.setColumnWidth(col_idx, cell_data.get('column_width', table.columnWidth(col_idx)))
                else:
                    item = QTableWidgetItem(cell_data)
                    item.setForeground(QColor(Qt.black))
                    item.setBackground(QColor(Qt.white))
                    table.setItem(row_idx, col_idx, item)
        step_end_time = time.time()
        print(f"Step 4 (Populate table with data) took {step_end_time - step_start_time:.4f} seconds")

        # Step 5: Set merged cells
        step_start_time = time.time()
        merged_cells = self.db_handler.get_merged_cells()
        if merged_cells["status"] == "success":
            merged_cells_data = merged_cells["data"]
            for cell in merged_cells_data:
                table.setSpan(int(cell['row']), int(cell['col']), int(cell['row_span']), int(cell['col_span']))
        step_end_time = time.time()
        print(f"Step 5 (Set merged cells) took {step_end_time - step_start_time:.4f} seconds")

        end_time = time.time()
        print(f"Total populate_table execution time: {end_time - start_time:.4f} seconds")

    def populate_table_with_default_data(self):
        table = self.table_widget.get_table()
        table.setRowCount(2)
        table.setColumnCount(11)
        for row in range(2):
            for col in range(11):
                item = QTableWidgetItem(f'({row}, {col})')
                item.setForeground(QColor(Qt.black))
                item.setBackground(QColor(Qt.white))
                table.setItem(row, col, item)
