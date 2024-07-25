from PySide6.QtWidgets import QTableWidgetItem, QFileDialog, QMessageBox
from PySide6.QtGui import QColor
from PySide6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
class TableHandler:
    def __init__(self, table_widget, db_handler):
        self.table_widget = table_widget
        self.db_handler = db_handler

    def save_data(self):
        table = self.table_widget.get_table()
        data = []
        for row in range(table.rowCount()):
            row_data = {}
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    font = item.font()
                    row_data[str(col)] = {
                        'text': item.text(),
                        'foreground': item.foreground().color().name(),
                        'background': item.background().color().name(),
                        'alignment': item.textAlignment(),
                        'font': {
                            'bold': font.bold(),
                            'size': font.pointSize()
                        },
                        'row_height': table.rowHeight(row),
                        'column_width': table.columnWidth(col)
                    }
                else:
                    row_data[str(col)] = {
                        'text': '',
                        'foreground': QColor(Qt.black).name(),
                        'background': QColor(Qt.white).name(),
                        'alignment': int(Qt.AlignLeft | Qt.AlignVCenter),
                        'font': {
                            'bold': False,
                            'size': 10
                        },
                        'row_height': table.rowHeight(row),
                        'column_width': table.columnWidth(col)
                    }
            data.append(row_data)

        merged_cells = []
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                if table.rowSpan(row, col) > 1 or table.columnSpan(row, col) > 1:
                    merged_cells.append({
                        'row': row,
                        'col': col,
                        'row_span': table.rowSpan(row, col),
                        'col_span': table.columnSpan(row, col)
                    })

        self.db_handler.save_table(data)
        self.db_handler.save_merged_cells(merged_cells)
        QMessageBox.information(self.table_widget, "保存成功", "表格数据已保存到数据库")

    def refresh_data(self):
        table = self.table_widget.get_table()
        if table.rowCount() > 0:
            table.removeRow(table.rowCount() - 1)
        self.load_table_data()

        table = self.table_widget.get_table()
        if table.rowCount() > 0:
            table.removeRow(table.rowCount() - 2)
        print("数据已刷新")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from PySide6.QtCore import Qt

    def export_to_excel(self):
        table = self.table_widget.get_table()
        path, _ = QFileDialog.getSaveFileName(self.table_widget, "导出为Excel", "", "Excel Files (*.xlsx)")
        if path:
            wb = Workbook()
            ws = wb.active

            # 遍历表格中的每个单元格
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        cell = ws.cell(row=row + 1, column=col + 1, value=item.text())

                        # 设置对齐方式
                        alignment = item.textAlignment()
                        align = Alignment(
                            horizontal='left' if alignment & Qt.AlignLeft else
                            'right' if alignment & Qt.AlignRight else
                            'center' if alignment & Qt.AlignHCenter else 'general',
                            vertical='top' if alignment & Qt.AlignTop else
                            'bottom' if alignment & Qt.AlignBottom else
                            'center' if alignment & Qt.AlignVCenter else 'center'
                        )
                        cell.alignment = align

                        # 设置字体
                        font = item.font()
                        cell.font = Font(
                            bold=font.bold(),
                            size=font.pointSize()
                        )

                        # 设置前景色和背景色
                        fg_color = item.foreground().color().name()
                        bg_color = item.background().color().name()
                        cell.fill = PatternFill(start_color=fg_color[1:], end_color=bg_color[1:], fill_type='solid')

                        # 设置单元格大小
                        ws.row_dimensions[row + 1].height = table.rowHeight(row)
                        ws.column_dimensions[get_column_letter(col + 1)].width = table.columnWidth(col) / 10
                    else:
                        ws.cell(row=row + 1, column=col + 1, value='')

            # 合并单元格
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    if table.rowSpan(row, col) > 1 or table.columnSpan(row, col) > 1:
                        ws.merge_cells(
                            start_row=row + 1, start_column=col + 1,
                            end_row=row + table.rowSpan(row, col),
                            end_column=col + table.columnSpan(row, col)
                        )

            wb.save(path)
            QMessageBox.information(self.table_widget, "导出成功", f"表格已成功导出到 {path}")

    def load_table_data(self):
        data = self.db_handler.get_table()
        if data:
            self.populate_table(data)
        else:
            self.populate_table_with_default_data()

    def populate_table(self, data):
        table = self.table_widget.get_table()
        table.clearContents()

        # Check if data is in the expected format
        if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
            raise ValueError("Data must be a list of dictionaries")

        row_count = len(data)
        col_count = max(len(row) for row in data) if row_count > 0 else 0
        table.setRowCount(row_count)
        table.setColumnCount(col_count)

        for row_idx, row_data in enumerate(data):
            if not isinstance(row_data, dict):
                raise ValueError(f"Row data at index {row_idx} is not a dictionary: {row_data}")

            for col_idx_str, cell_data in row_data.items():
                try:
                    col_idx = int(col_idx_str)  # Convert column index back to integer
                except ValueError:
                    continue  # Skip keys that cannot be converted to integers

                if isinstance(cell_data, dict):
                    item = QTableWidgetItem(cell_data.get('text', ''))
                    item.setForeground(QColor(cell_data.get('foreground', QColor(Qt.black).name())))
                    item.setBackground(QColor(cell_data.get('background', QColor(Qt.white).name())))
                    item.setTextAlignment(cell_data.get('alignment', int(Qt.AlignLeft | Qt.AlignVCenter)))
                    font = item.font()
                    font.setBold(cell_data.get('font', {}).get('bold', False))
                    font.setPointSize(cell_data.get('font', {}).get('size', 10))
                    item.setFont(font)
                    table.setItem(row_idx, col_idx, item)
                    table.setRowHeight(row_idx, cell_data.get('row_height', table.rowHeight(row_idx)))
                    table.setColumnWidth(col_idx, cell_data.get('column_width', table.columnWidth(col_idx)))
                else:
                    item = QTableWidgetItem(cell_data)
                    item.setForeground(QColor(Qt.black))
                    item.setBackground(QColor(Qt.white))
                    table.setItem(row_idx, col_idx, item)

        merged_cells = self.db_handler.get_merged_cells()
        for cell in merged_cells:
            table.setSpan(cell['row'], cell['col'], cell['row_span'], cell['col_span'])

    def populate_table_with_default_data(self):
        table = self.table_widget.get_table()
        table.setRowCount(2)
        table.setColumnCount(11)
        for row in range(2):
            for col in range(11):
                item = QTableWidgetItem(f'({row}, {col})')
                item.setForeground(QColor(Qt.black))
                item.setBackground(QColor(Qt.white))
                table.setItem(row, col, item)
